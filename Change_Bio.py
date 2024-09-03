import tweepy
import time
import datetime
import openpyxl
from openpyxl import load_workbook

def changebio():
    #On recupere la liste des bio
    """
    chemin_bio="C:/Users/Administrator/Desktop/OneDrive/SHARED GFP/Developpement/Python/TwitterVaiot/template_influencer.xlsx"

    wb = load_workbook(chemin_bio)
    index = 0
    for index in range(len(wb.sheetnames)):
        if wb.sheetnames[index] == "Data":
            break
    wb.active = index
    sheet = wb.active

    liste_bio=[]
    indiceExcel=2
    while str(sheet['G' + str(indiceExcel)].value) != "None":
        liste_bio.append(sheet['G' + str(indiceExcel)].value)
        indiceExcel=indiceExcel+1

    print(liste_bio)
    """

    liste_bio=["Learn about Blockchain, Web 3.0, NFT and Cryptocurrencies news in Africa with Web3africa: https://web3africa.news/"]

    #on recupere les api associées au projet et on change la bio. il peut y avoir moins de bio que d'api, plusieurs api auront la même bio alors
    chemin_api="C:/Users/Administrator/Desktop/OneDrive/SHARED GFP/Developpement/Python/TwitterVaiot/Project/APIs_Project.xlsx"
    wb = load_workbook(chemin_api)
    index = 0
    for index in range(len(wb.sheetnames)):
        if wb.sheetnames[index] == "Data":
            break
    wb.active = index
    sheet = wb.active

    indice_bio=0
    compteur_bio=0
    indiceExcel=2
    while str(sheet['C' + str(indiceExcel)].value) != "None":
        print("Name API: " + str(sheet['A' + str(indiceExcel)].value))
        try:
            api_key = sheet['C' + str(indiceExcel)].value
            api_key_secret = sheet['D' + str(indiceExcel)].value
            bearer = sheet['E' + str(indiceExcel)].value
            access_token = sheet['F' + str(indiceExcel)].value
            access_token_secret = sheet['G' + str(indiceExcel)].value
            proxyapi=sheet['J' + str(indiceExcel)].value
            print(proxyapi)
            autherticator = tweepy.OAuthHandler(api_key, api_key_secret)
            autherticator.set_access_token(access_token, access_token_secret)
            proxyapi = "http://" + str(proxyapi)
            api = tweepy.API(autherticator, proxy=proxyapi)

            #on recupere la bio de la liste dont l'indice est indice_bio modulo longueur de la liste, et on increment ensuite indice_bio +1
            q,indice_bio=divmod(indice_bio, len(liste_bio))
            bio = liste_bio[indice_bio] + " (" + str(compteur_bio) + ")"
            #api.update_profile(url="https://web3africa.news/", description=bio)
            api.update_profile(url="", description=bio)
            print("Done")
            compteur_bio=compteur_bio+1
            indice_bio = indice_bio + 1
            sheet['L' + str(indiceExcel)] = bio
        except Exception as e:
            print('error', e)
            # return e
        else:
            a = 2
        time.sleep(40)
        indiceExcel=indiceExcel+1

    wb.save(chemin_api)

changebio()
time.sleep(5)

def changer_banner():

    banner="image8.jpg"

    #on recupere les api associées au projet
    chemin_api="C:/Users/Administrator/Desktop/OneDrive/SHARED GFP/Developpement/Python/TwitterVaiot/Project/APIs_Project.xlsx"
    wb = load_workbook(chemin_api)
    index = 0
    for index in range(len(wb.sheetnames)):
        if wb.sheetnames[index] == "Data":
            break
    wb.active = index
    sheet = wb.active

    indiceExcel=2
    while str(sheet['C' + str(indiceExcel)].value) != "None":
        print("Name API: " + str(sheet['A' + str(indiceExcel)].value))
        try:
            api_key = sheet['C' + str(indiceExcel)].value
            api_key_secret = sheet['D' + str(indiceExcel)].value
            bearer = sheet['E' + str(indiceExcel)].value
            access_token = sheet['F' + str(indiceExcel)].value
            access_token_secret = sheet['G' + str(indiceExcel)].value
            proxyapi = sheet['J' + str(indiceExcel)].value
            print(proxyapi)
            autherticator = tweepy.OAuthHandler(api_key, api_key_secret)
            autherticator.set_access_token(access_token, access_token_secret)
            proxyapi = "http://" + str(proxyapi)
            api = tweepy.API(autherticator, proxy=proxyapi)

            api.update_profile_banner(filename=banner)

        except Exception as e:
            print('error', e)
            # return e
        else:
            a = 2
        time.sleep(40)
        indiceExcel=indiceExcel+1

#changer_banner()
#time.sleep(5)
def changer_photoprofil():
    #mettre le chemin si image pas dans le meme dossier que ce fichier python
    profil = "image23.jpg"

    # on recupere les api associées au projet
    chemin_api = "C:/Users/Administrator/Desktop/OneDrive/SHARED GFP/Developpement/Python/TwitterVaiot/Project/APIs_Project.xlsx"
    wb = load_workbook(chemin_api)
    index = 0
    for index in range(len(wb.sheetnames)):
        if wb.sheetnames[index] == "Data":
            break
    wb.active = index
    sheet = wb.active

    indiceExcel = 2
    while str(sheet['C' + str(indiceExcel)].value) != "None":
        print("Name API: " + str(sheet['A' + str(indiceExcel)].value))
        try:
            api_key = sheet['C' + str(indiceExcel)].value
            api_key_secret = sheet['D' + str(indiceExcel)].value
            bearer = sheet['E' + str(indiceExcel)].value
            access_token = sheet['F' + str(indiceExcel)].value
            access_token_secret = sheet['G' + str(indiceExcel)].value
            proxyapi = sheet['J' + str(indiceExcel)].value
            print(proxyapi)
            autherticator = tweepy.OAuthHandler(api_key, api_key_secret)
            autherticator.set_access_token(access_token, access_token_secret)
            proxyapi = "http://" + str(proxyapi)
            api = tweepy.API(autherticator, proxy=proxyapi)

            api.update_profile_image(filename=profil)

        except Exception as e:
            print('error', e)
            # return e
        else:
            a = 2
        time.sleep(40)
        indiceExcel = indiceExcel + 1

#changer_photoprofil()